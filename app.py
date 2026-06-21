import os
import json
import shutil
import uuid
import time
from datetime import datetime, timedelta, timezone
from flask import (Flask, render_template, request, redirect, url_for,
                   flash, jsonify, send_from_directory, session, abort)
from flask_login import (LoginManager, login_user, logout_user,
                          login_required, current_user)
from flask_wtf.csrf import CSRFProtect
from werkzeug.utils import secure_filename
from werkzeug.exceptions import HTTPException
from functools import wraps

from config import Config
from models import (db, User, College, StudyYear, Semester, Course, Lecture,
                    LectureFile, Test, Question, TestResult, StudentEnrollment,
                    Announcement, UniversityLogo, BackupLog, StudentReward,
                    StudentTracking, TeacherAssignment, ForumPost, ForumReply, ChatMessage, LiveSession, WhiteboardStroke)
from utils.decorators import admin_required, teacher_required, student_required, manage_required
from utils.helpers import (perform_backup, get_college_name, generate_ai_questions)
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_limiter.errors import RateLimitExceeded





app = Flask(__name__)
app.config.from_object(Config)

# ===== الأمن السيبراني المعزز (Rate Limiting) =====
limiter = Limiter(get_remote_address, app=app, default_limits=['200 per minute'], storage_uri='memory://')

@app.errorhandler(RateLimitExceeded)
def ratelimit_handler(e):
    lang = session.get('lang', 'ar')
    msg = 'لقد أرسلت طلبات كثيرة جداً. يرجى الانتظار قليلاً.' if lang == 'ar' else 'Too many requests. Please slow down.'
    return f"<h1 style='text-align:center; padding:50px; color:red; font-family:Cairo, sans-serif;'>{msg}</h1>", 429

app.secret_key = os.environ.get('SECRET_KEY', 'ghyath-university-platform-2026-super-secret-key-x9k2m')

@app.errorhandler(Exception)
def handle_all_errors(e):
    import traceback
    if isinstance(e, HTTPException):
        return e
    traceback.print_exc()
    return "<h1 style='text-align:center;padding:50px;color:red;'>حدث خطأ غير متوقع في الخادم، يرجى المحاولة لاحقاً</h1>", 500

csrf = CSRFProtect(app)
# ===== هيدرز الأمن السيبراني =====
@app.after_request
def set_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    return response

db.init_app(app)


app.jinja_env.globals['now'] = datetime.now
app.jinja_env.globals['timedelta'] = timedelta
@app.template_filter('count_items')
def count_items(obj):
    try:
        return obj.count()
    except (TypeError, AttributeError):
        try:
            return len(obj)
        except TypeError:
            return 0

login_manager = LoginManager(app)
login_manager.login_view = 'login'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

login_manager.login_message = ''
login_manager.login_message_category = 'warning'
login_manager.session_protection = 'strong'

def guest_forbidden(message='عليك التسجيل بالمنصة أولاً للوصول إلى هذا القسم'):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if current_user.is_authenticated and current_user.is_guest:
                flash(message, 'warning')
                return redirect(url_for('index'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

@app.before_request
def check_authentication():
    if request.endpoint == 'static' or request.path.startswith('/static/') or request.path == '/favicon.ico':
        return None
    allowed_routes = ['index', 'login', 'logout', 'change_language', 'guest_login', 'guest_exit']
    if request.endpoint in allowed_routes:
        return None
    if current_user.is_authenticated:
        return None
    if request.method == 'GET':
        session['next_url'] = request.url
    flash('عليك تسجيل الدخول إلى المنصة أولاً', 'warning')
    return redirect(url_for('login'))

@login_manager.unauthorized_handler
def unauthorized_access():
    if request.method == 'GET':
        session['next_url'] = request.url
    flash('عليك تسجيل الدخول إلى المنصة أولاً', 'warning')
    return redirect(url_for('login'))

def get_lang():
    return session.get('lang', 'ar')

def check_student_enrolled(course_id):
    if not current_user.is_authenticated or not current_user.is_student:
        return False
    enrollment = StudentEnrollment.query.filter_by(user_id=current_user.id, course_id=course_id).first()
    return enrollment is not None and enrollment.is_approved

def require_enrollment(course_id):
    if not check_student_enrolled(course_id):
        flash('عليك التسجيل والموافقة عليه أولاً لتتمكن من الوصول إليه', 'warning')
        return redirect(url_for('student_courses'))
    return None

def get_logo():
    return UniversityLogo.query.filter_by(is_active=True).order_by(UniversityLogo.created_at.desc()).first()

def allowed_file(filename, allowed_extensions):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions

def save_file(file, folder, allowed_extensions):
    if file and allowed_file(file.filename, allowed_extensions):
        os.makedirs(folder, exist_ok=True)
        ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
        unique_name = "{}_{}.{}".format(uuid.uuid4().hex[:8], datetime.now().strftime('%Y%m%d%H%M%S'), ext)
        filepath = os.path.join(folder, unique_name)
        file.save(filepath)
        return unique_name, filepath
    return None, None
    # ===== قاعدة بيانات مؤقتة في الذاكرة لحساب الطلبات =====

@app.context_processor
def inject_globals():
    def nav():
        from utils.helpers import get_navigation_structure
        return get_navigation_structure()

    def translate_test_type(test_type, lang='ar'):
        translations = {
            'mcq': {'ar': 'اختيار من متعدد', 'en': 'MCQ'},
            'tf': {'ar': 'صح أو خطأ', 'en': 'T/F'},
            'essay': {'ar': 'مقالي', 'en': 'Essay'},
            'exercise': {'ar': 'تمارين', 'en': 'Exercises'},
            'homework': {'ar': 'واجبات', 'en': 'Homework'},
            'quiz': {'ar': 'مذاكرة', 'en': 'Quiz'},
            'discussion': {'ar': 'مناقشة', 'en': 'Discussion'},
            'media': {'ar': 'صوت وصورة', 'en': 'Media'},
            'forum': {'ar': 'منتدى', 'en': 'Forum'},
            'online': {'ar': 'عبر الإنترنت', 'en': 'Online'},
        }
        return translations.get(test_type, {}).get(lang, test_type or '')

    return dict(get_navigation_structure=nav, translate_test_type=translate_test_type)

@app.route('/')
def index():
    lang = get_lang()
    logo = get_logo()
    colleges = College.query.filter_by(parent_id=None).filter(College.is_active == True).order_by(College.sort_order).all()
    total_students = User.query.filter_by(user_type='student', is_active=True).count()
    total_teachers = User.query.filter_by(user_type='teacher', is_active=True).count()
    total_courses = Course.query.filter_by(is_active=True).count()
    return render_template('index.html', lang=lang, logo=logo, colleges=colleges, total_students=total_students, total_teachers=total_teachers, total_courses=total_courses)

@app.route('/change-language/<lang_code>')
def change_language(lang_code):
    if lang_code in ['ar', 'en']:
        session['lang'] = lang_code
    return redirect(request.referrer or url_for('index'))

@limiter.limit('5 per minute')
@app.route('/login', methods=['GET', 'POST'])
def login():
    lang = get_lang()
    logo = get_logo()
    if current_user.is_authenticated:
        return redirect(url_for('index'))

    failed_attempts = session.get('failed_login_attempts', 0)
    last_attempt_time = session.get('last_login_attempt', 0)

    if failed_attempts >= 5:
        wait_seconds = 300
        elapsed = datetime.now().timestamp() - last_attempt_time
        if elapsed < wait_seconds:
            remaining = int(wait_seconds - elapsed)
            flash('تم تجاوز عدد المحاولات المسموحة. حاول مرة أخرى بعد {} ثانية'.format(remaining), 'danger')
            return render_template('login.html', lang=lang, logo=logo, locked_out=True, remaining=remaining)
        else:
            session['failed_login_attempts'] = 0

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        user = User.query.filter_by(username=username, is_active=True).first()
        if user and user.check_password(password):
            session.pop('failed_login_attempts', None)
            session.pop('last_login_attempt', None)
            user.last_login = datetime.now(timezone.utc)
            if user.login_count is None:
                user.login_count = 1
            else:
                user.login_count += 1
            db.session.commit()
            login_user(user, remember=True)
            role_names = {'admin': 'مدير المنصة', 'teacher': 'مدرس المقرر', 'student': 'طالب', 'guest': 'ضيف زائر'}
            welcome_msg = role_names.get(user.user_type, 'مستخدم')
            flash('مرحباً بك يا {} {} ({}) في منصة جامعة الجزيرة الخاصة'.format(user.first_name, user.second_name, welcome_msg), 'success')
            next_url = session.pop('next_url', None)
            return redirect(next_url or url_for('index'))
        else:
            session['failed_login_attempts'] = failed_attempts + 1
            session['last_login_attempt'] = datetime.now().timestamp()
            remaining_attempts = 5 - (failed_attempts + 1)
            if remaining_attempts > 0:
                flash('اسم المستخدم أو كلمة المرور غير صحيحة (متبقي {} محاولات)'.format(remaining_attempts), 'danger')
            else:
                flash('تم تجاوز عدد المحاولات. يرجى الانتظار 5 دقائق', 'danger')
    return render_template('login.html', lang=lang, logo=logo)

@app.route('/login/guest')
def guest_login():
    guest_user = User.query.filter_by(user_type='guest', username='guest_visitor').first()
    if not guest_user:
        guest_user = User(first_name='ضيف', second_name='زائر', username='guest_visitor', user_type='guest', is_active=True)
        guest_user.set_password('guest_secure_password_2026')
        db.session.add(guest_user)
        db.session.commit()
    login_user(guest_user, remember=False)
    flash('مرحباً بك كضيف زائر! يمكنك التجول في المنصة بحرية، ولكن للوصول للمحتوى التعليمي عليك التسجيل.', 'info')
    return redirect(url_for('index'))

@app.route('/guest-exit')
@login_required
def guest_exit():
    logout_user()
    flash('يرجى تسجيل الدخول باستخدام حسابك الفعلي كطالب أو مدرس للاستفادة من كافة الميزات', 'success')
    return redirect(url_for('login'))

@app.route('/logout')
@login_required
def logout():
    name = current_user.full_name
    role_names = {'admin': 'مدير المنصة', 'teacher': 'مدرس المقرر', 'student': 'طالب', 'guest': 'ضيف زائر'}
    role = role_names.get(current_user.user_type, 'مستخدم')
    logout_user()
    flash('مع السلامة يا {} ({})، تم تسجيل الخروج بنجاح'.format(name, role), 'info')
    return redirect(url_for('index'))

@app.route('/college/<int:id>')
def college_detail(id):
    lang = get_lang()
    logo = get_logo()
    college = College.query.get_or_404(id)
    return render_template('colleges/detail.html', lang=lang, logo=logo, college=college)

@app.route('/college/<int:id>/add', methods=['GET', 'POST'])
@admin_required
def add_college(id):
    lang = get_lang()
    logo = get_logo()
    parent = College.query.get_or_404(id)
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        name_en = request.form.get('name_en', '').strip()
        if not name_en and name:
            name_en = translate_to_english(name)
        description = request.form.get('description', '').strip()
        if name:
            college = College(name=name, name_en=name_en, description=description, parent_id=parent.id, level=parent.level + 1)
            db.session.add(college)
            db.session.commit()
            flash('تم إضافة الكلية/القسم بنجاح', 'success')
            return redirect(url_for('college_detail', id=parent.id))
    return render_template('colleges/add.html', lang=lang, logo=logo, parent=parent)

@app.route('/college/<int:id>/edit', methods=['GET', 'POST'])
@admin_required
def edit_college(id):
    lang = get_lang()
    logo = get_logo()
    college = College.query.get_or_404(id)
    if request.method == 'POST':
        college.name = request.form.get('name', '').strip()
        college.name_en = request.form.get('name_en', '').strip()
        if not college.name_en and college.name:
            college.name_en = translate_to_english(college.name)
        college.description = request.form.get('description', '').strip()
        db.session.commit()
        flash('تم تعديل الكلية/القسم بنجاح', 'success')
        return redirect(url_for('college_detail', id=college.parent_id if college.parent_id else college.id))
    return render_template('colleges/edit.html', lang=lang, logo=logo, college=college)

@app.route('/college/<int:id>/delete', methods=['POST'])
@admin_required
def delete_college(id):
    college = College.query.get_or_404(id)
    parent_id = college.parent_id
    try:
        db.session.delete(college)
        db.session.commit()
        flash('تم حذف الكلية/القسم بنجاح', 'success')
    except:
        db.session.rollback()
        flash('لا يمكن حذف هذا العنصر لوجود بيانات مرتبطة', 'danger')
    if parent_id:
        return redirect(url_for('college_detail', id=parent_id))
    return redirect(url_for('index'))

@app.route('/year/<int:id>')
def year_detail(id):
    lang = get_lang()
    logo = get_logo()
    year = StudyYear.query.get_or_404(id)
    semesters = Semester.query.filter_by(year_id=id).order_by(Semester.sort_order).all()
    return render_template('colleges/years.html', lang=lang, logo=logo, year=year, semesters=semesters)

@app.route('/year/<int:id>/add', methods=['GET', 'POST'])
@admin_required
def add_year(id):
    lang = get_lang()
    logo = get_logo()
    college = College.query.get_or_404(id)
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        name_en = request.form.get('name_en', '').strip()
        if not name_en and name:
            name_en = translate_to_english(name)
        description = request.form.get('description', '').strip()
        if name:
            year = StudyYear(college_id=college.id, name=name, name_en=name_en, description=description)
            db.session.add(year)
            db.session.commit()
            for sn, sen, st, so in [('الفصل الدراسي الأول','First Semester','first',1),('الفصل الدراسي الثاني','Second Semester','second',2),('الفصل الدراسي الصيفي','Summer Semester','summer',3)]:
                db.session.add(Semester(year_id=year.id, name=sn, name_en=sen, semester_type=st, sort_order=so))
            db.session.commit()
            flash('تم إضافة السنة الدراسية مع الفصول الثلاثة بنجاح', 'success')
            return redirect(url_for('college_detail', id=college.id))
    return render_template('colleges/add_year.html', lang=lang, logo=logo, college=college)

@app.route('/year/<int:id>/edit', methods=['GET', 'POST'])
@admin_required
def edit_year(id):
    lang = get_lang()
    logo = get_logo()
    year = StudyYear.query.get_or_404(id)
    if request.method == 'POST':
        year.name = request.form.get('name', '').strip()
        year.name_en = request.form.get('name_en', '').strip()
        if not year.name_en and year.name:
            year.name_en = translate_to_english(year.name)
        year.description = request.form.get('description', '').strip()
        db.session.commit()
        flash('تم تعديل السنة الدراسية بنجاح', 'success')
        return redirect(url_for('year_detail', id=year.id))
    return render_template('colleges/edit_year.html', lang=lang, logo=logo, year=year)

@app.route('/year/<int:id>/delete', methods=['POST'])
@admin_required
def delete_year(id):
    year = StudyYear.query.get_or_404(id)
    college_id = year.college_id
    try:
        db.session.delete(year)
        db.session.commit()
        flash('تم حذف السنة الدراسية بنجاح', 'success')
    except:
        db.session.rollback()
        flash('لا يمكن حذف هذه السنة لوجود بيانات مرتبطة', 'danger')
    return redirect(url_for('college_detail', id=college_id))

@app.route('/year/<int:id>/fix-semesters', methods=['POST'])
@admin_required
def fix_year_semesters(id):
    year = StudyYear.query.get_or_404(id)
    if Semester.query.filter_by(year_id=id).count() == 0:
        for sn, sen, st, so in [('الفصل الدراسي الأول','First Semester','first',1),('الفصل الدراسي الثاني','Second Semester','second',2),('الفصل الدراسي الصيفي','Summer Semester','summer',3)]:
            db.session.add(Semester(year_id=year.id, name=sn, name_en=sen, semester_type=st, sort_order=so))
        db.session.commit()
        flash('تم إنشاء الفصول الثلاثة بنجاح', 'success')
    else:
        flash('الفصول موجودة مسبقاً', 'info')
    return redirect(url_for('year_detail', id=id))

@app.route('/semester/<int:id>')
def semester_detail(id):
    lang = get_lang()
    logo = get_logo()
    semester = Semester.query.get_or_404(id)
    return render_template('colleges/semesters.html', lang=lang, logo=logo, semester=semester)

@app.route('/semester/<int:id>/edit', methods=['GET', 'POST'])
@admin_required
def edit_semester(id):
    lang = get_lang()
    logo = get_logo()
    semester = Semester.query.get_or_404(id)
    if request.method == 'POST':
        semester.name = request.form.get('name', '').strip()
        semester.name_en = request.form.get('name_en', '').strip()
        semester.description = request.form.get('description', '').strip()
        db.session.commit()
        flash('تم تعديل الفصل الدراسي بنجاح', 'success')
        return redirect(url_for('semester_detail', id=semester.id))
    return render_template('colleges/edit_semester.html', lang=lang, logo=logo, semester=semester)

@app.route('/semester/<int:id>/delete', methods=['POST'])
@admin_required
def delete_semester(id):
    semester = Semester.query.get_or_404(id)
    year_id = semester.year_id
    try:
        db.session.delete(semester)
        db.session.commit()
        flash('تم حذف الفصل الدراسي بنجاح', 'success')
    except:
        db.session.rollback()
        flash('لا يمكن حذف هذا الفصل لوجود بيانات مرتبطة', 'danger')
    return redirect(url_for('year_detail', id=year_id))

@app.route('/course/<int:id>')
def course_detail(id):
    lang = get_lang()
    logo = get_logo()
    course = Course.query.get_or_404(id)
    return render_template('colleges/courses.html', lang=lang, logo=logo, course=course)

@app.route('/course/<int:id>/add', methods=['GET', 'POST'])
@teacher_required
def add_course(id):
    lang = get_lang()
    logo = get_logo()
    semester = Semester.query.get_or_404(id)
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        name_en = request.form.get('name_en', '').strip()
        if not name_en and name:
            name_en = translate_to_english(name)
        code = request.form.get('code', '').strip()
        description = request.form.get('description', '').strip()
        if name:
            course = Course(semester_id=semester.id, name=name, name_en=name_en, code=code, description=description, teacher_id=current_user.id)
            db.session.add(course)
            db.session.commit()
            flash('تم إضافة المقرر الدراسي بنجاح', 'success')
            return redirect(url_for('semester_detail', id=semester.id))
    return render_template('colleges/add_course.html', lang=lang, logo=logo, semester=semester)

@app.route('/course/<int:id>/edit', methods=['GET', 'POST'])
@teacher_required
def edit_course(id):
    lang = get_lang()
    logo = get_logo()
    course = Course.query.get_or_404(id)
    if request.method == 'POST':
        course.name = request.form.get('name', '').strip()
        course.name_en = request.form.get('name_en', '').strip()
        if not course.name_en and course.name:
            course.name_en = translate_to_english(course.name)
        course.code = request.form.get('code', '').strip()
        course.description = request.form.get('description', '').strip()
        if current_user.is_admin:
            tid = request.form.get('teacher_id', type=int)
            if tid:
                course.teacher_id = tid
        db.session.commit()
        flash('تم تعديل المقرر الدراسي بنجاح', 'success')
        return redirect(url_for('course_detail', id=course.id))
    teachers = User.query.filter_by(user_type='teacher', is_active=True).all() if current_user.is_admin else []
    return render_template('colleges/edit_course.html', lang=lang, logo=logo, course=course, teachers=teachers)

@app.route('/course/<int:id>/delete', methods=['POST'])
@admin_required
def delete_course(id):
    course = Course.query.get_or_404(id)
    semester_id = course.semester_id
    try:
        db.session.delete(course)
        db.session.commit()
        flash('تم حذف المقرر الدراسي بنجاح', 'success')
    except:
        db.session.rollback()
        flash('لا يمكن حذف هذا المقرر لوجود بيانات مرتبطة', 'danger')
    return redirect(url_for('semester_detail', id=semester_id))

@app.route('/lecture/<int:id>')
@guest_forbidden('عليك التسجيل بالمنصة أولاً لمعاينة المحاضرات')
def lecture_detail(id):
    lang = get_lang()
    logo = get_logo()
    lecture = Lecture.query.get_or_404(id)
    return render_template('colleges/lectures.html', lang=lang, logo=logo, lecture=lecture)

@app.route('/lecture/<int:id>/add', methods=['GET', 'POST'])
@teacher_required
def add_lecture(id):
    lang = get_lang()
    logo = get_logo()
    course = Course.query.get_or_404(id)
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        name_en = request.form.get('name_en', '').strip()
        if not name_en and name:
            name_en = translate_to_english(name)
        description = request.form.get('description', '').strip()
        if name:
            lecture = Lecture(course_id=course.id, name=name, name_en=name_en, description=description)
            db.session.add(lecture)
            db.session.commit()
            flash('تم إضافة المحاضرة الدراسية بنجاح', 'success')
            return redirect(url_for('course_detail', id=course.id))
    return render_template('colleges/add_lecture.html', lang=lang, logo=logo, course=course)

@app.route('/lecture/<int:id>/edit', methods=['GET', 'POST'])
@teacher_required
def edit_lecture(id):
    lang = get_lang()
    logo = get_logo()
    lecture = Lecture.query.get_or_404(id)
    if request.method == 'POST':
        lecture.name = request.form.get('name', '').strip()
        lecture.name_en = request.form.get('name_en', '').strip()
        if not lecture.name_en and lecture.name:
            lecture.name_en = translate_to_english(lecture.name)
        lecture.description = request.form.get('description', '').strip()
        db.session.commit()
        flash('تم تعديل المحاضرة الدراسية بنجاح', 'success')
        return redirect(url_for('lecture_detail', id=lecture.id))
    return render_template('colleges/edit_lecture.html', lang=lang, logo=logo, lecture=lecture)

@app.route('/lecture/<int:id>/delete', methods=['POST'])
@admin_required
def delete_lecture(id):
    lecture = Lecture.query.get_or_404(id)
    course_id = lecture.course_id
    try:
        db.session.delete(lecture)
        db.session.commit()
        flash('تم حذف المحاضرة الدراسية بنجاح', 'success')
    except:
        db.session.rollback()
        flash('لا يمكن حذف هذه المحاضرة لوجود بيانات مرتبطة', 'danger')
    return redirect(url_for('course_detail', id=course_id))

@app.route('/lecture/<int:id>/upload', methods=['GET', 'POST'])
@teacher_required
def upload_lecture(id):
    lang = get_lang()
    logo = get_logo()
    lecture = Lecture.query.get_or_404(id)
    if request.method == 'POST':
        file_type = request.form.get('file_type', '')
        external_link = request.form.get('external_link', '').strip()
        file_obj = request.files.get('file')
        saved = False
        if file_type == 'link':
            if external_link:
                lf = LectureFile(lecture_id=lecture.id, file_type='link', file_name='رابط خارجي', external_link=external_link)
                db.session.add(lf)
                saved = True
        elif file_obj:
            allowed = {'ppt': ['ppt','pptx'], 'video': ['mp4','avi','mkv','webm'], 'audio': ['mp3','wav','ogg'], 'pdf': ['pdf']}
            exts = allowed.get(file_type, [])
            if exts:
                fname, fpath = save_file(file_obj, app.config['LECTURE_FOLDER'], exts)
                if fname:
                    lf = LectureFile(lecture_id=lecture.id, file_type=file_type, file_name=fname, file_path=fpath, file_size=os.path.getsize(fpath) if os.path.exists(fpath) else 0)
                    db.session.add(lf)
                    saved = True
        if saved:
            db.session.commit()
            flash('تم رفع المحاضرة بنجاح', 'success')
        else:
            flash('فشل في رفع المحاضرة', 'danger')
        return redirect(url_for('lecture_detail', id=lecture.id))
    return render_template('colleges/upload.html', lang=lang, logo=logo, lecture=lecture)

@app.route('/lecture-file/<int:id>/delete', methods=['POST'])
@teacher_required
def delete_lecture_file(id):
    lf = LectureFile.query.get_or_404(id)
    lecture_id = lf.lecture_id
    if lf.file_path and os.path.exists(lf.file_path):
        os.remove(lf.file_path)
    db.session.delete(lf)
    db.session.commit()
    flash('تم حذف الملف بنجاح', 'success')
    return redirect(url_for('lecture_detail', id=lecture_id))

@app.route('/lecture-file/<int:id>/preview')
@login_required
@guest_forbidden('عليك التسجيل بالمنصة أولاً لمعاينة الملفات')
def preview_lecture_file(id):
    lf = LectureFile.query.get_or_404(id)
    if current_user.is_student and lf.lecture:
        if not check_student_enrolled(lf.lecture.course_id):
            flash('عليك التسجيل في هذا المقرر أولاً', 'warning')
            return redirect(url_for('student_courses'))
    if lf.file_type == 'link' and lf.external_link:
        return redirect(lf.external_link)
    if lf.file_path and os.path.exists(lf.file_path):
        return send_from_directory(os.path.dirname(lf.file_path), os.path.basename(lf.file_path))
    abort(404)

@app.route('/lecture-file/<int:id>/download')
@login_required
@guest_forbidden('عليك التسجيل بالمنصة أولاً لتحميل الملفات')
def download_lecture_file(id):
    lf = LectureFile.query.get_or_404(id)
    if current_user.is_student and lf.lecture:
        if not check_student_enrolled(lf.lecture.course_id):
            flash('عليك التسجيل في هذا المقرر أولاً', 'warning')
            return redirect(url_for('student_courses'))
    if lf.file_path and os.path.exists(lf.file_path):
        return send_from_directory(os.path.dirname(lf.file_path), os.path.basename(lf.file_path), as_attachment=True, download_name=lf.file_name or 'lecture_file')
    abort(404)

@app.route('/admin/dashboard')
@admin_required
def admin_dashboard():
    lang = get_lang()
    logo = get_logo()
    total_users = User.query.count()
    total_students = User.query.filter_by(user_type='student').count()
    total_teachers = User.query.filter_by(user_type='teacher').count()
    total_courses = Course.query.count()
    total_tests = Test.query.count()
    total_lectures = Lecture.query.count()
    recent_users = User.query.order_by(User.created_at.desc()).limit(5).all()
    return render_template('admin/dashboard.html', lang=lang, logo=logo, total_users=total_users, total_students=total_students, total_teachers=total_teachers, total_courses=total_courses, total_tests=total_tests, total_lectures=total_lectures, recent_users=recent_users)

@app.route('/admin/users')
@admin_required
def admin_users():
    lang = get_lang()
    logo = get_logo()
    page = request.args.get('page', 1, type=int)
    per_page = 50
    user_type_filter = request.args.get('type', '')
    search = request.args.get('search', '')
    query = User.query
    if user_type_filter:
        query = query.filter_by(user_type=user_type_filter)
    if search:
        query = query.filter(db.or_(User.first_name.contains(search), User.second_name.contains(search), User.username.contains(search), User.mobile.contains(search)))
    users = query.order_by(User.created_at.desc()).paginate(page=page, per_page=per_page)
    return render_template('admin/users.html', lang=lang, logo=logo, users=users, user_type_filter=user_type_filter, search=search)

@app.route('/admin/users/add', methods=['GET', 'POST'])
@admin_required
def admin_add_user():
    lang = get_lang()
    logo = get_logo()
    all_colleges = College.query.filter_by(is_active=True).order_by(College.sort_order).all()
    all_years = StudyYear.query.order_by(StudyYear.sort_order).all()
    if request.method == 'POST':
        number = request.form.get('number', type=int)
        first_name = request.form.get('first_name', '').strip()
        second_name = request.form.get('second_name', '').strip()
        mobile = request.form.get('mobile', '').strip()
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        user_type = request.form.get('user_type', '').strip()
        college_id = request.form.get('college_id', type=int)
        study_year = request.form.get('study_year', '').strip()
        if not all([first_name, second_name, username, password, user_type]):
            flash('يرجى ملء جميع الحقول المطلوبة', 'danger')
            return redirect(url_for('admin_add_user'))
        if User.query.filter_by(username=username).first():
            flash('اسم المستخدم موجود مسبقاً', 'danger')
            return redirect(url_for('admin_add_user'))
        user = User(number=number, first_name=first_name, second_name=second_name, mobile=mobile, username=username, user_type=user_type, college_id=college_id, study_year=study_year)
        user.set_password(password)
        db.session.add(user)
        db.session.flush()
        if user_type == 'teacher':
            additional = request.form.getlist('additional_colleges')
            for cid_str in additional:
                cid_int = int(cid_str)
                if cid_int != (college_id or 0):
                    ay_key = 'ay_{}'.format(cid_str)
                    ay_id = request.form.get(ay_key, type=int)
                    ta = TeacherAssignment(teacher_id=user.id, college_id=cid_int, study_year_id=ay_id)
                    db.session.add(ta)
        db.session.commit()
        flash('تم إضافة المستخدم بنجاح', 'success')
        return redirect(url_for('admin_users'))
    return render_template('admin/add_user.html', lang=lang, logo=logo, all_colleges=all_colleges, all_years=all_years)

@app.route('/admin/users/upload-excel', methods=['GET', 'POST'])
@admin_required
def admin_upload_excel():
    lang = get_lang()
    logo = get_logo()
    if request.method == 'POST':
        file = request.files.get('excel_file')
        if file and allowed_file(file.filename, ['xlsx', 'xls']):
            try:
                import openpyxl
                filepath = os.path.join(app.config['EXCEL_FOLDER'], secure_filename(file.filename))
                file.save(filepath)
                wb = openpyxl.load_workbook(filepath)
                ws = wb.active
                added = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] and row[4] and row[5]:
                        if not User.query.filter_by(username=str(row[4])).first():
                            user = User(number=int(row[0]) if row[0] else None, first_name=str(row[1] or ''), second_name=str(row[2] or ''), mobile=str(row[3] or ''), username=str(row[4]), user_type=str(row[6] or 'student'))
                            user.set_password(str(row[5]))
                            db.session.add(user)
                            added += 1
                db.session.commit()
                flash('تم استيراد {} مستخدم'.format(added), 'success')
            except Exception as e:
                flash('خطأ: {}'.format(str(e)), 'danger')
        else:
            flash('يرجى رفع ملف xlsx', 'danger')
        return redirect(url_for('admin_users'))
    return render_template('admin/upload_excel.html', lang=lang, logo=logo)

@app.route('/admin/users/download-template')
@admin_required
def admin_download_template():
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'نموذج'
    for col, h in enumerate(['الرقم','الاسم الأول','الاسم الثاني','الموبايل','اسم المستخدم','كلمة المرور','النوع'], 1):
        ws.cell(row=1, column=col, value=h)
    for col, v in enumerate([1001,'محمد','أحمد','09XXXXXXXX','mohammed','pass123','student'], 1):
        ws.cell(row=2, column=col, value=v)
    fp = os.path.join(app.config['EXCEL_FOLDER'], 'template_users.xlsx')
    wb.save(fp)
    return send_from_directory(app.config['EXCEL_FOLDER'], 'template_users.xlsx', as_attachment=True, download_name='نموذج.xlsx')

@app.route('/admin/users/edit/<int:id>', methods=['GET', 'POST'])
@admin_required
def admin_edit_user(id):
    lang = get_lang()
    logo = get_logo()
    user = User.query.get_or_404(id)
    all_colleges = College.query.filter_by(is_active=True).order_by(College.sort_order).all()
    all_years = StudyYear.query.order_by(StudyYear.sort_order).all()
    assignments = TeacherAssignment.query.filter_by(teacher_id=user.id).all() if user.is_teacher else []
    if request.method == 'POST':
        user.number = request.form.get('number', type=int)
        user.first_name = request.form.get('first_name', '').strip()
        user.second_name = request.form.get('second_name', '').strip()
        user.mobile = request.form.get('mobile', '').strip()
        user.username = request.form.get('username', '').strip()
        user.user_type = request.form.get('user_type', '').strip()
        user.college_id = request.form.get('college_id', type=int)
        user.study_year = request.form.get('study_year', '').strip()
        np = request.form.get('new_password', '').strip()
        if np:
            user.set_password(np)
        if user.is_teacher:
            TeacherAssignment.query.filter_by(teacher_id=user.id).delete()
            additional = request.form.getlist('additional_colleges')
            for cid_str in additional:
                cid_int = int(cid_str)
                if cid_int != (user.college_id or 0):
                    ay_key = 'ay_{}'.format(cid_str)
                    ay_id = request.form.get(ay_key, type=int)
                    ta = TeacherAssignment(teacher_id=user.id, college_id=cid_int, study_year_id=ay_id)
                    db.session.add(ta)
        db.session.commit()
        flash('تم تعديل المستخدم بنجاح', 'success')
        return redirect(url_for('admin_users'))
    return render_template('admin/edit_user.html', lang=lang, logo=logo, user=user, all_colleges=all_colleges, all_years=all_years, assignments=assignments)

@app.route('/admin/users/delete/<int:id>', methods=['POST'])
@admin_required
def admin_delete_user(id):
    user = User.query.get_or_404(id)
    if user.id == current_user.id:
        flash('لا يمكنك حذف حسابك', 'danger')
        return redirect(url_for('admin_users'))
    try:
        db.session.delete(user)
        db.session.commit()
        flash('تم الحذف', 'success')
    except:
        db.session.rollback()
        flash('لا يمكن الحذف', 'danger')
    return redirect(url_for('admin_users'))

@app.route('/admin/statistics')
@admin_required
def admin_statistics():
    lang = get_lang()
    logo = get_logo()
    page = request.args.get('page', 1, type=int)
    per_page = 50
    stats = {'total_users': User.query.count(), 'students': User.query.filter_by(user_type='student').count(), 'teachers': User.query.filter_by(user_type='teacher').count(), 'courses': Course.query.count(), 'tests': Test.query.count(), 'lectures': Lecture.query.count()}
    registered_users = User.query.order_by(User.created_at.desc()).paginate(page=page, per_page=per_page)
    test_type_stats = {}
    for t in Test.query.all():
        type_name = t.test_type or 'غير محدد'
        test_type_stats[type_name] = test_type_stats.get(type_name, 0) + 1
    colleges_stats = []
    for c in College.query.filter_by(parent_id=None, is_active=True).all():
        depts_count = College.query.filter_by(parent_id=c.id).count()
        sub_college_ids = [c.id] + [d.id for d in College.query.filter_by(parent_id=c.id).all()]
        years = StudyYear.query.filter(StudyYear.college_id.in_(sub_college_ids)).all()
        yrs_count = len(years)
        crs_count = 0
        for y in years:
            semesters = Semester.query.filter_by(year_id=y.id).all()
            for s in semesters:
                crs_count += Course.query.filter_by(semester_id=s.id).count()
        colleges_stats.append({'name': c.name, 'departments': depts_count, 'years': yrs_count, 'courses': crs_count})
    return render_template('admin/statistics.html', lang=lang, logo=logo, stats=stats, registered_users=registered_users, test_type_stats=test_type_stats, colleges_stats=colleges_stats)

@app.route('/admin/tests')
@admin_required
def admin_tests():
    lang = get_lang()
    logo = get_logo()
    tf = request.args.get('type', '')
    query = Test.query
    if tf:
        query = query.filter_by(test_type=tf)
    tests = query.order_by(Test.created_at.desc()).all()
    return render_template('admin/tests.html', lang=lang, logo=logo, tests=tests, test_type_filter=tf)

@app.route('/admin/tests/add', methods=['GET', 'POST'])
@teacher_required
def admin_add_test():
    lang = get_lang()
    logo = get_logo()
    if request.method == 'POST':
        course_id = request.form.get('course_id', type=int)
        name = request.form.get('name', '').strip()
        test_type = request.form.get('test_type', '').strip()
        total_mark = request.form.get('total_mark', type=float)
        per_question_mark = request.form.get('per_question_mark', type=float)
        duration = request.form.get('duration_minutes', type=int)
        deadline_str = request.form.get('deadline', '')
        deadline = None
        if deadline_str:
            try:
                deadline = datetime.strptime(deadline_str, '%Y-%m-%dT%H:%M')
            except ValueError:
                flash('صيغة التاريخ غير صحيحة', 'warning')
        if not all([course_id, name, test_type]):
            flash('يرجى ملء الحقول', 'danger')
            return redirect(url_for('admin_add_test'))
        test = Test(course_id=course_id, name=name, test_type=test_type, total_mark=total_mark or 100, per_question_mark=per_question_mark or 1, duration_minutes=duration or 60, deadline=deadline, created_by=current_user.id)
        db.session.add(test)
        db.session.commit()
        flash('تم إضافة الاختبار', 'success')
        if current_user.is_admin:
            return redirect(url_for('admin_tests'))
        else:
            return redirect(url_for('teacher_tests'))
    courses = Course.query.all() if current_user.is_admin else Course.query.filter_by(teacher_id=current_user.id).all()
    return render_template('admin/add_test.html', lang=lang, logo=logo, courses=courses)

@app.route('/admin/tests/edit/<int:id>', methods=['GET', 'POST'])
@teacher_required
def admin_edit_test(id):
    lang = get_lang()
    logo = get_logo()
    test = Test.query.get_or_404(id)
    if request.method == 'POST':
        test.name = request.form.get('name', '').strip()
        test.test_type = request.form.get('test_type', '').strip()
        test.total_mark = request.form.get('total_mark', type=float) or 100
        test.per_question_mark = request.form.get('per_question_mark', type=float) or 1
        test.duration_minutes = request.form.get('duration_minutes', type=int) or 60
        ds = request.form.get('deadline', '')
        test.deadline = datetime.strptime(ds, '%Y-%m-%dT%H:%M') if ds else None
        db.session.commit()
        flash('تم التعديل', 'success')
        if current_user.is_admin:
            return redirect(url_for('admin_tests'))
        else:
            return redirect(url_for('teacher_tests'))
    courses = Course.query.all() if current_user.is_admin else Course.query.filter_by(teacher_id=current_user.id).all()
    return render_template('admin/edit_test.html', lang=lang, logo=logo, test=test, courses=courses)

@app.route('/admin/tests/delete/<int:id>', methods=['POST'])
@admin_required
def admin_delete_test(id):
    test = Test.query.get_or_404(id)
    try:
        db.session.delete(test)
        db.session.commit()
        flash('تم الحذف', 'success')
    except:
        db.session.rollback()
        flash('لا يمكن الحذف', 'danger')
    return redirect(url_for('admin_tests'))

@app.route('/admin/tests/toggle/<int:id>', methods=['POST'])
@teacher_required
def toggle_test_status(id):
    test = Test.query.get_or_404(id)
    if current_user.is_teacher and not current_user.is_admin:
        course = Course.query.get(test.course_id)
        if course.teacher_id != current_user.id:
            abort(403)
    test.is_active = not test.is_active
    db.session.commit()
    status_text = "مفعل" if test.is_active else "معطل"
    flash('تم {} الاختبار بنجاح'.format(status_text), 'success')
    if current_user.is_admin:
        return redirect(url_for('admin_tests'))
    return redirect(url_for('teacher_tests'))

@app.route('/admin/tests/<int:id>/questions')
@teacher_required
def test_questions(id):
    lang = get_lang()
    logo = get_logo()
    test = Test.query.get_or_404(id)
    questions = test.questions.order_by(Question.sort_order).all()
    if current_user.is_admin:
        back_url = url_for('admin_tests')
    else:
        back_url = url_for('teacher_tests')
    return render_template('admin/test_questions.html', lang=lang, logo=logo, test=test, questions=questions, back_url=back_url)

@app.route('/admin/tests/<int:id>/questions/add', methods=['GET', 'POST'])
@teacher_required
def add_test_question(id):
    lang = get_lang()
    logo = get_logo()
    test = Test.query.get_or_404(id)
    if request.method == 'POST':
        qt = request.form.get('question_text', '').strip()
        qtype = request.form.get('question_type', '').strip()
        opts = request.form.getlist('options')
        ca = request.form.get('correct_answer', '').strip()
        marks = request.form.get('marks', type=float) or test.per_question_mark
        if qt:
            q = Question(test_id=test.id, question_text=qt, question_type=qtype, options=json.dumps(opts, ensure_ascii=False) if opts else None, correct_answer=ca, marks=marks)
            db.session.add(q)
            db.session.commit()
            flash('تم الإضافة', 'success')
            return redirect(url_for('test_questions', id=test.id))
    if current_user.is_admin:
        back_url = url_for('admin_tests')
    else:
        back_url = url_for('teacher_tests')
    return render_template('admin/add_question.html', lang=lang, logo=logo, test=test, back_url=back_url)

@limiter.limit('10 per minute')
@app.route('/admin/tests/<int:test_id>/questions/generate-ai', methods=['POST'])
@teacher_required
def generate_ai_test_questions(test_id):
    test = Test.query.get_or_404(test_id)
    topic = request.form.get('topic', '').strip()
    num = request.form.get('num_questions', type=int) or 5

    if num > 15:
        num = 15

    if not topic:
        flash('يرجى إدخال موضوع لتوليد الأسئلة', 'warning')
        return redirect(url_for('test_questions', id=test_id))

    try:
        qs = generate_ai_questions(topic, num, test.test_type)

        if not qs or len(qs) == 0:
            flash('لم يتم توليد أي أسئلة. يرجى المحاولة مع موضوع مختلف', 'danger')
            return redirect(url_for('test_questions', id=test_id))

        added = 0
        for qd in qs:
            if not isinstance(qd, dict):
                continue

            question_text = str(qd.get('question_text', '')).strip()
            if not question_text:
                continue

            options = qd.get('options')
            if isinstance(options, list):
                options = json.dumps(options, ensure_ascii=False)
            elif isinstance(options, str):
                try:
                    parsed = json.loads(options)
                    options = json.dumps(parsed, ensure_ascii=False)
                except (json.JSONDecodeError, TypeError):
                    options = json.dumps([options], ensure_ascii=False) if options else None
            else:
                options = None

            try:
                marks = float(qd.get('marks', test.per_question_mark or 1))
            except (ValueError, TypeError):
                marks = float(test.per_question_mark or 1)

            question = Question(
                test_id=test.id,
                question_text=question_text,
                question_type=str(qd.get('question_type', 'essay')),
                options=options,
                correct_answer=str(qd.get('correct_answer', '')),
                marks=marks
            )
            db.session.add(question)
            added += 1

        if added > 0:
            db.session.commit()
            flash('✅ تم إنشاء {} سؤال ذكي بنجاح حول الموضوع: {}'.format(added, topic), 'success')
        else:
            db.session.rollback()
            flash('لم يتم إنشاء أي أسئلة صالحة من المصدر', 'danger')

    except Exception as e:
        db.session.rollback()
        error_msg = str(e)[:300]
        print(f"AI Generation Error: {error_msg}")
        flash('⚠️ حدث خطأ أثناء توليد الأسئلة: {}'.format(error_msg), 'danger')

    return redirect(url_for('test_questions', id=test_id))

@app.route('/admin/tests/<int:test_id>/questions/delete/<int:q_id>', methods=['POST'])
@teacher_required
def delete_test_question(test_id, q_id):
    q = Question.query.get_or_404(q_id)
    test = Test.query.get(test_id)
    if test is None:
        abort(404)

    if current_user.is_teacher and not current_user.is_admin:
        course = Course.query.get(test.course_id)
        if course.teacher_id != current_user.id:
            abort(403)

    try:
        db.session.delete(q)
        db.session.commit()
        flash('تم حذف السؤال بنجاح', 'success')
    except:
        db.session.rollback()
        flash('لا يمكن حذف هذا السؤال', 'danger')

    return redirect(url_for('test_questions', id=test_id))

@app.route('/admin/tests/<int:id>/results')
@teacher_required
def test_results(id):
    lang = get_lang()
    logo = get_logo()
    test = Test.query.get_or_404(id)
    results = test.results.order_by(TestResult.submitted_at.desc()).all()
    if current_user.is_admin:
        back_url = url_for('admin_tests')
    else:
        back_url = url_for('teacher_tests')
    return render_template('admin/test_results.html', lang=lang, logo=logo, test=test, results=results, back_url=back_url)

@app.route('/admin/logo', methods=['GET', 'POST'])
@admin_required
def admin_logo():
    lang = get_lang()
    logo = get_logo()
    logos = UniversityLogo.query.order_by(UniversityLogo.created_at.desc()).all()
    if request.method == 'POST':
        file = request.files.get('logo_file')
        if file and allowed_file(file.filename, ['png','jpg','jpeg','svg','gif','webp']):
            fn, fp = save_file(file, app.config['LOGO_FOLDER'], ['png','jpg','jpeg','svg','gif','webp'])
            if fn:
                UniversityLogo.query.update({UniversityLogo.is_active: False})
                db.session.add(UniversityLogo(file_name=fn, file_path=fp, is_active=True))
                db.session.commit()
                flash('تم رفع الشعار', 'success')
        else:
            flash('صيغة غير مدعومة', 'danger')
        return redirect(url_for('admin_logo'))
    return render_template('admin/logo.html', lang=lang, logo=logo, logos=logos)

@app.route('/admin/logo/activate/<int:id>', methods=['POST'])
@admin_required
def activate_logo(id):
    logo = UniversityLogo.query.get_or_404(id)
    UniversityLogo.query.update({UniversityLogo.is_active: False})
    logo.is_active = True
    db.session.commit()
    flash('تم التفعيل', 'success')
    return redirect(url_for('admin_logo'))

@app.route('/admin/backup')
@admin_required
def admin_backup():
    lang = get_lang()
    logo = get_logo()
    logs = BackupLog.query.order_by(BackupLog.created_at.desc()).limit(20).all()
    return render_template('admin/backup.html', lang=lang, logo=logo, logs=logs)

@app.route('/admin/backup/trigger', methods=['POST'])
@admin_required
def trigger_backup():
    result = perform_backup()
    if result:
        flash('تم النسخ الاحتياطي', 'success')
    else:
        flash('فشل', 'danger')
    return redirect(url_for('admin_backup'))

@app.route('/admin/backup/download/<int:id>')
@admin_required
def download_backup(id):
    log = BackupLog.query.get_or_404(id)
    if log.file_path and os.path.exists(log.file_path):
        return send_from_directory(os.path.dirname(log.file_path), os.path.basename(log.file_path), as_attachment=True)
    abort(404)

@app.route('/admin/enrollments')
@admin_required
def admin_enrollments():
    lang = get_lang()
    logo = get_logo()
    enrollments = StudentEnrollment.query.order_by(StudentEnrollment.is_approved.asc(), StudentEnrollment.enrolled_at.desc()).all()
    return render_template('admin/enrollments.html', lang=lang, logo=logo, enrollments=enrollments)

@app.route('/admin/enrollments/approve/<int:id>', methods=['POST'])
@admin_required
def approve_enrollment(id):
    enrollment = StudentEnrollment.query.get_or_404(id)
    enrollment.is_approved = True
    db.session.commit()
    flash('تمت الموافقة على التسجيل بنجاح', 'success')
    return redirect(url_for('admin_enrollments'))

@app.route('/admin/enrollments/upload-excel', methods=['POST'])
@admin_required
def upload_enrollment_excel():
    file = request.files.get('excel_file')
    if file and allowed_file(file.filename, ['xlsx', 'xls']):
        try:
            import openpyxl
            filepath = os.path.join(app.config['EXCEL_FOLDER'], secure_filename(file.filename))
            file.save(filepath)
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            approved_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                student_number = row[0]
                course_name = row[5]
                if student_number and course_name:
                    user = User.query.filter_by(number=int(student_number), user_type='student').first()
                    course = Course.query.filter_by(name=str(course_name).strip()).first()
                    if user and course:
                        enrollment = StudentEnrollment.query.filter_by(user_id=user.id, course_id=course.id).first()
                        if enrollment and not enrollment.is_approved:
                            enrollment.is_approved = True
                            approved_count += 1
                        elif not enrollment:
                            new_enr = StudentEnrollment(user_id=user.id, course_id=course.id, is_approved=True)
                            db.session.add(new_enr)
                            approved_count += 1
            db.session.commit()
            flash('تمت الموافقة على {} طلب تسجيل من ملف الإكسل'.format(approved_count), 'success')
        except Exception as e:
            flash('خطأ في قراءة الملف: {}'.format(str(e)), 'danger')
    else:
        flash('يرجى رفع ملف xlsx', 'danger')
    return redirect(url_for('admin_enrollments'))

@app.route('/teacher/dashboard')
@teacher_required
def teacher_dashboard():
    lang = get_lang()
    logo = get_logo()
    courses = Course.query.filter_by(teacher_id=current_user.id, is_active=True).all()
    total_students = sum(c.enrollments.count() for c in courses)
    total_tests = sum(c.tests.count() for c in courses)
    total_lectures = sum(c.lectures.count() for c in courses)
    announcements_sent = Announcement.query.filter_by(sender_id=current_user.id).count()
    primary_college = College.query.get(current_user.college_id) if current_user.college_id else None
    teacher_assignments = TeacherAssignment.query.filter_by(teacher_id=current_user.id).all()
    assigned_colleges = []
    seen_ids = set()
    if primary_college and primary_college.id not in seen_ids:
        assigned_colleges.append({'college': primary_college, 'is_primary': True, 'study_year': None})
        seen_ids.add(primary_college.id)
    for ta in teacher_assignments:
        col = College.query.get(ta.college_id)
        if col and col.id not in seen_ids:
            yr = StudyYear.query.get(ta.study_year_id) if ta.study_year_id else None
            assigned_colleges.append({'college': col, 'is_primary': False, 'study_year': yr})
            seen_ids.add(col.id)
    return render_template('teacher/dashboard.html', lang=lang, logo=logo, courses=courses, total_students=total_students, total_tests=total_tests, total_lectures=total_lectures, announcements_sent=announcements_sent, primary_college=primary_college, assigned_colleges=assigned_colleges)

@app.route('/teacher/courses')
@teacher_required
def teacher_courses():
    lang = get_lang()
    logo = get_logo()
    courses = Course.query.filter_by(teacher_id=current_user.id).order_by(Course.created_at.desc()).all()
    return render_template('teacher/courses.html', lang=lang, logo=logo, courses=courses)

@app.route('/teacher/lectures')
@teacher_required
def teacher_lectures():
    lang = get_lang()
    logo = get_logo()
    courses = Course.query.filter_by(teacher_id=current_user.id).all()
    lectures = [l for c in courses for l in c.lectures.all()]
    return render_template('teacher/lectures.html', lang=lang, logo=logo, lectures=lectures)

@app.route('/teacher/tests')
@teacher_required
def teacher_tests():
    lang = get_lang()
    logo = get_logo()
    courses = Course.query.filter_by(teacher_id=current_user.id).all()
    course_ids = [c.id for c in courses]
    query = Test.query.filter(Test.course_id.in_(course_ids))

    test_type_filter = request.args.get('type', '').strip()
    if test_type_filter:
        query = query.filter_by(test_type=test_type_filter)

    search = request.args.get('search', '').strip()
    if search:
        query = query.filter(Test.name.contains(search))

    course_filter = request.args.get('course_id', 0, type=int)
    if course_filter:
        query = query.filter_by(course_id=course_filter)

    status_filter = request.args.get('status', '').strip()
    if status_filter == 'active':
        query = query.filter_by(is_active=True)
    elif status_filter == 'inactive':
        query = query.filter_by(is_active=False)
    elif status_filter == 'expired':
        query = query.filter(Test.deadline != None, Test.deadline < datetime.utcnow())

    tests = query.order_by(Test.created_at.desc()).all()
    total_submissions = sum(t.results.count() for t in tests) if tests else 0
    total_questions = sum(t.questions.count() for t in tests) if tests else 0
    active_tests = sum(1 for t in tests if t.is_active)
    expired_tests = sum(1 for t in tests if t.deadline and t.deadline < datetime.utcnow())
    avg_submissions = round(total_submissions / len(tests), 1) if tests else 0

    return render_template('teacher/tests.html', lang=lang, logo=logo, tests=tests,
        total_submissions=total_submissions, total_questions=total_questions,
        test_type_filter=test_type_filter, search=search, course_filter=course_filter,
        status_filter=status_filter, courses=courses, active_tests=active_tests,
        expired_tests=expired_tests, avg_submissions=avg_submissions)

@app.route('/teacher/test/<int:test_id>/submissions')
@teacher_required
def teacher_view_submissions(test_id):
    lang = get_lang()
    logo = get_logo()
    test = Test.query.get_or_404(test_id)
    results = test.results.order_by(TestResult.submitted_at.desc()).all()
    return render_template('admin/test_results.html', lang=lang, logo=logo, test=test, results=results, back_url=url_for('teacher_tests'))

@app.route('/teacher/students-tracking')
@teacher_required
def teacher_students_tracking():
    lang = get_lang()
    logo = get_logo()
    courses = Course.query.filter_by(teacher_id=current_user.id).all()
    course_id = request.args.get('course_id', type=int)
    selected_course = None
    students_data = []
    page = request.args.get('page', 1, type=int)
    per_page = 25

    if course_id:
        selected_course = Course.query.get_or_404(course_id)
        all_enrollments = StudentEnrollment.query.filter_by(course_id=course_id, is_approved=True).all()
        total_students = len(all_enrollments)
        start = (page - 1) * per_page
        end = start + per_page
        paginated_enrollments = all_enrollments[start:end]

        for e in paginated_enrollments:
            s = e.user
            if not s:
                continue

            all_results = TestResult.query.filter_by(user_id=s.id).join(Test).filter(Test.course_id == course_id).all()

            total_tests = len(all_results)
            overall_avg = sum(r.percentage for r in all_results) / total_tests if total_tests > 0 else 0
            best_score = max((r.percentage for r in all_results), default=0)

            quiz_results = [r for r in all_results if r.test and r.test.test_type in ['مذاكرة', 'quiz', 'مذاكرات']]
            muthakarat_taken = len(quiz_results)
            muthakarat_avg = sum(r.percentage for r in quiz_results) / muthakarat_taken if muthakarat_taken > 0 else 0

            assign_results = [r for r in all_results if r.test and r.test.test_type in ['وظيفة', 'assignment', 'تمرين', 'واجب']]
            assignments_taken = len(assign_results)
            assignments_avg = sum(r.percentage for r in assign_results) / assignments_taken if assignments_taken > 0 else 0

            rewards_count = StudentReward.query.filter_by(user_id=s.id).count()
            tracking = StudentTracking.query.filter_by(user_id=s.id, course_id=course_id).order_by(StudentTracking.created_at.desc()).limit(5).all()

            has_activity = len(tracking) > 0

            if total_tests == 0:
                if not has_activity:
                    ai_eval = "👻 غائب تماماً (لا أثر له في المقرر)"
                    ai_tip = "يجب التواصل مع المرشد الأكاديمي فوراً لمعرفة سبب الانقطاع التام"
                    ai_class = "dark"
                else:
                    ai_eval = "👁️ متفرج (يتابع المحتوى لكنه لا يختبر)"
                    ai_tip = "الطالب حاضر إلكترونياً لكنه يتهرب من التقييم، يرجى تحفيزه"
                    ai_class = "secondary"
            else:
                if overall_avg >= 90:
                    ai_eval = "🌟 متفوق (أداء استثنائي ومرتفع)"
                    ai_tip = "يُنصح بتكليفه بمهام قيادية أو مساعدة زملائه كمعيد"
                    ai_class = "success"
                elif overall_avg >= 75:
                    ai_eval = "✅ جيد جداً (أداء متميز ومستقر)"
                    ai_tip = "يحتاج لتحفيز بسيط للوصول لمستوى التفوق والامتياز"
                    ai_class = "primary"
                elif overall_avg >= 60:
                    ai_eval = "⚡ جيد (يحتاج متابعة وتركيز أكبر)"
                    ai_tip = "يُنصح بتكثيف التمارين والمراجعة لضمان عدم الانخفاض"
                    ai_class = "warning"
                else:
                    ai_eval = "⚠️ ضعيف (يحتاج تدخل عاجل ودعم أكاديمي)"
                    ai_tip = "يجب عقد جلسة دعم فردي لمعرفة العقبات ومساعدته قبل فوات الأوان"
                    ai_class = "danger"

            students_data.append({
                'student': s,
                'overall_avg': round(overall_avg, 1),
                'total_tests': total_tests,
                'best_score': round(best_score, 1),
                'muthakarat_taken': muthakarat_taken,
                'muthakarat_avg': round(muthakarat_avg, 1),
                'assignments_taken': assignments_taken,
                'assignments_avg': round(assignments_avg, 1),
                'rewards': rewards_count,
                'tracking': tracking,
                'ai_eval': ai_eval,
                'ai_tip': ai_tip,
                'ai_class': ai_class
            })
        return render_template('teacher/students_tracking.html', lang=lang, logo=logo, courses=courses, selected_course=selected_course, students_data=students_data, course_id=course_id, page=page, per_page=per_page, total_students=total_students)

    return render_template('teacher/students_tracking.html', lang=lang, logo=logo, courses=courses, selected_course=selected_course, students_data=students_data, course_id=course_id, page=1, per_page=per_page, total_students=0)

@app.route('/teacher/send-announcement', methods=['GET', 'POST'])
@teacher_required
def teacher_send_announcement():
    lang = get_lang()
    logo = get_logo()
    courses = Course.query.filter_by(teacher_id=current_user.id).all()
    if request.method == 'POST':
        course_id = request.form.get('course_id', type=int)
        title = request.form.get('title', '').strip()
        message = request.form.get('message', '').strip()
        msg_type = request.form.get('msg_type', 'general')
        priority = request.form.get('priority', 'normal')
        ds = request.form.get('deadline', '')
        deadline = datetime.strptime(ds, '%Y-%m-%dT%H:%M') if ds else None
        if course_id and message:
            sent = 0
            for e in StudentEnrollment.query.filter_by(course_id=course_id).all():
                db.session.add(Announcement(sender_id=current_user.id, receiver_id=e.user_id, course_id=course_id, title=title, message=message, msg_type=msg_type, priority=priority, deadline=deadline))
                sent += 1
            db.session.commit()
            flash('تم الإرسال لـ {} طالب'.format(sent), 'success')
            return redirect(url_for('teacher_send_announcement'))
    return render_template('teacher/send_announcement.html', lang=lang, logo=logo, courses=courses)

@app.route('/teacher/student-rewards', methods=['GET', 'POST'])
@teacher_required
def teacher_student_rewards():
    lang = get_lang()
    logo = get_logo()
    courses = Course.query.filter_by(teacher_id=current_user.id).all()
    if request.method == 'POST':
        uid = request.form.get('user_id', type=int)
        rtype = request.form.get('reward_type', '').strip()
        rname = request.form.get('reward_name', '').strip()
        desc = request.form.get('description', '').strip()
        pts = request.form.get('points', type=int) or 0
        if uid and rname:
            db.session.add(StudentReward(user_id=uid, reward_type=rtype, reward_name=rname, description=desc, points=pts))
            db.session.commit()
            flash('تم منح الجائزة', 'success')
            return redirect(url_for('teacher_student_rewards'))
    rewards = StudentReward.query.order_by(StudentReward.created_at.desc()).limit(50).all()
    return render_template('teacher/student_rewards.html', lang=lang, logo=logo, courses=courses, rewards=rewards)

@app.route('/auto-enroll')
@login_required
def auto_enroll():
    if not current_user.is_student:
        flash('هذه الميزة متاحة للطلاب فقط', 'warning')
        return redirect(url_for('index'))

    student = current_user
    added = 0
    student_college = College.query.get(student.college_id) if student.college_id else None
    if not student_college:
        flash('لم يتم تحديد كليتك بعد. تواصل مع الإدارة', 'warning')
        return redirect(url_for('student_dashboard'))

    college_ids = set()
    college_ids.add(student_college.id)
    if student_college.parent_id:
        college_ids.add(student_college.parent_id)
    children = College.query.filter_by(parent_id=student_college.id).all()
    for child in children:
        college_ids.add(child.id)
    if student_college.parent_id:
        siblings = College.query.filter_by(parent_id=student_college.parent_id).all()
        for s in siblings:
            college_ids.add(s.id)

    year_ids = [y.id for y in StudyYear.query.filter(StudyYear.college_id.in_(college_ids)).all()]
    if student.study_year:
        try:
            target_year = int(student.study_year)
            year_ids = [yid for yid in year_ids if yid == target_year]
        except (ValueError, TypeError):
            pass
    semester_ids = [s.id for s in Semester.query.filter(Semester.year_id.in_(year_ids)).all()]
    courses = Course.query.filter(Course.semester_id.in_(semester_ids), Course.is_active == True).all()

    for course in courses:
        existing = StudentEnrollment.query.filter_by(user_id=student.id, course_id=course.id).first()
        if not existing:
            db.session.add(StudentEnrollment(user_id=student.id, course_id=course.id, is_approved=True))
            added += 1
    if added > 0:
        db.session.commit()
        flash('تم تسجيلك تلقائيا في {} مقرر'.format(added), 'success')
    else:
        flash('مسجل في جميع المقررات بالفعل', 'info')
    return redirect(url_for('student_dashboard'))

LOCAL_TIMEZONE_OFFSET = 3

def get_local_now():
    return datetime.utcnow() + timedelta(hours=LOCAL_TIMEZONE_OFFSET)

def update_session_statuses(course_id):
    now = get_local_now()
    upcoming = LiveSession.query.filter(LiveSession.course_id == course_id, LiveSession.status == 'upcoming', LiveSession.start_time <= now).all()
    for s in upcoming:
        s.status = 'live'
    live = LiveSession.query.filter(LiveSession.course_id == course_id, LiveSession.status == 'live', LiveSession.end_time <= now).all()
    for s in live:
        s.status = 'finished'
    db.session.commit()

@app.route('/course/<int:course_id>/track-focus', methods=['POST'])
@login_required
def track_focus(course_id):
    if not current_user.is_student:
        return jsonify({'success': False})
    data = request.get_json()
    is_focused = data.get('focused', True)
    action_type = 'session_focus' if is_focused else 'session_unfocus'
    last_track = StudentTracking.query.filter_by(user_id=current_user.id, course_id=course_id).order_by(StudentTracking.created_at.desc()).first()
    if not last_track or last_track.action_type != action_type:
        track = StudentTracking(user_id=current_user.id, course_id=course_id, action_type=action_type, action_details='طالب يتابع الجلسة' if is_focused else 'طالب غادر شاشة الجلسة')
        db.session.add(track)
        db.session.commit()
    return jsonify({'success': True})

@app.route('/course/<int:course_id>/get-focus-statuses', methods=['GET'])
@login_required
def get_focus_statuses(course_id):
    if not (current_user.is_teacher or current_user.is_admin):
        return jsonify({})
    enrollments = StudentEnrollment.query.filter_by(course_id=course_id, is_approved=True).all()
    statuses = {}
    for e in enrollments:
        last_track = StudentTracking.query.filter_by(user_id=e.user_id, course_id=course_id).order_by(StudentTracking.created_at.desc()).first()
        statuses[str(e.user_id)] = 'unfocused' if (last_track and last_track.action_type == 'session_unfocus') else 'focused'
    return jsonify(statuses)

@app.route('/course/<int:course_id>/virtual-class')
@login_required
@guest_forbidden('عليك التسجيل بالمنصة أولاً للدخول إلى الغرفة الافتراضية')
def virtual_classroom(course_id):
    lang = get_lang()
    logo = get_logo()
    course = Course.query.get_or_404(course_id)
    if current_user.is_student and not check_student_enrolled(course_id):
        flash('عليك التسجيل والموافقة عليه أولاً للدخول للغرفة الافتراضية', 'danger')
        return redirect(url_for('student_courses'))
    update_session_statuses(course_id)
    posts = ForumPost.query.filter_by(course_id=course_id).order_by(ForumPost.is_pinned.desc(), ForumPost.created_at.desc()).all()
    chat_messages = ChatMessage.query.filter_by(course_id=course_id).order_by(ChatMessage.created_at.asc()).limit(50).all()
    enrolled_students = []
    if current_user.is_teacher or current_user.is_admin:
        enrollments = StudentEnrollment.query.filter_by(course_id=course_id, is_approved=True).all()
        for e in enrollments:
            user = User.query.get(e.user_id)
            if user:
                enrolled_students.append(user)
    teacher = User.query.get(course.teacher_id)
    chat_user_ids = set(msg.user_id for msg in chat_messages)
    chat_users = {u.id: u for u in User.query.filter(User.id.in_(chat_user_ids)).all()} if chat_user_ids else {}
    sessions = LiveSession.query.filter_by(course_id=course_id).order_by(LiveSession.start_time.desc()).all()
    next_session = LiveSession.query.filter(LiveSession.course_id == course_id, LiveSession.status.in_(['upcoming', 'live'])).order_by(LiveSession.start_time.asc()).first()
    next_session_start_str = (next_session.start_time + timedelta(hours=LOCAL_TIMEZONE_OFFSET)).strftime('%Y-%m-%dT%H:%M:%S') if next_session else ''
    next_session_end_str = (next_session.end_time + timedelta(hours=LOCAL_TIMEZONE_OFFSET)).strftime('%Y-%m-%dT%H:%M:%S') if next_session else ''
    return render_template('virtual_classroom.html', lang=lang, logo=logo, course=course, posts=posts, chat_messages=chat_messages, enrolled_students=enrolled_students, teacher=teacher, chat_users=chat_users, sessions=sessions, next_session=next_session, next_session_start_str=next_session_start_str, next_session_end_str=next_session_end_str)

@app.route('/course/<int:course_id>/class/save-stroke', methods=['POST'])
@teacher_required
def save_stroke(course_id):
    data = request.get_json()
    stroke = data.get('stroke')
    if stroke:
        db.session.add(WhiteboardStroke(course_id=course_id, user_id=current_user.id, stroke_data=json.dumps(stroke)))
        db.session.commit()
        last = WhiteboardStroke.query.order_by(WhiteboardStroke.id.desc()).first()
        return jsonify({'success': True, 'id': last.id})
    return jsonify({'success': False})

@app.route('/course/<int:course_id>/class/get-strokes', methods=['GET'])
@login_required
def get_strokes(course_id):
    update_session_statuses(course_id)
    last_id = request.args.get('last_id', 0, type=int)
    new_strokes = WhiteboardStroke.query.filter(WhiteboardStroke.course_id == course_id, WhiteboardStroke.id > last_id).order_by(WhiteboardStroke.id.asc()).limit(500).all()
    next_session = LiveSession.query.filter(LiveSession.course_id == course_id, LiveSession.status.in_(['upcoming', 'live'])).order_by(LiveSession.start_time.asc()).first()
    session_info = None
    if next_session:
        session_info = {'id': next_session.id, 'title': next_session.title, 'status': next_session.status, 'start_str': (next_session.start_time + timedelta(hours=LOCAL_TIMEZONE_OFFSET)).strftime('%Y-%m-%dT%H:%M:%S'), 'end_str': (next_session.end_time + timedelta(hours=LOCAL_TIMEZONE_OFFSET)).strftime('%Y-%m-%dT%H:%M:%S')}
    return jsonify({'strokes': [json.loads(s.stroke_data) for s in new_strokes], 'last_id': new_strokes[-1].id if new_strokes else last_id, 'session': session_info})

@app.route('/course/<int:course_id>/class/clear-board-db', methods=['POST'])
@teacher_required
def clear_board_db(course_id):
    cutoff = datetime.utcnow() - timedelta(hours=24)
    old = WhiteboardStroke.query.filter(WhiteboardStroke.course_id == course_id, WhiteboardStroke.created_at < cutoff).all()
    for s in old:
        db.session.delete(s)
    db.session.commit()
    return jsonify({'success': True, 'deleted': len(old)})

@app.route('/course/<int:course_id>/class/upload-board-file', methods=['POST'])
@teacher_required
def upload_board_file(course_id):
    file = request.files.get('board_file')
    if file and allowed_file(file.filename, ['png', 'jpg', 'jpeg', 'pdf']):
        board_folder = os.path.join(app.static_folder, 'board_uploads')
        os.makedirs(board_folder, exist_ok=True)
        fname, fpath = save_file(file, board_folder, ['png', 'jpg', 'jpeg', 'pdf'])
        if fname:
            file_url = url_for('static', filename='board_uploads/' + fname)
            return jsonify({'success': True, 'file_url': file_url})
    return jsonify({'success': False})

@app.route('/course/<int:course_id>/class/schedule', methods=['POST'])
@teacher_required
def schedule_session(course_id):
    start_str = request.form.get('start_time', '')
    end_str = request.form.get('end_time', '')
    custom_title = request.form.get('session_title', '').strip()
    if start_str and end_str:
        course = Course.query.get(course_id)
        course_name = course.name if course else "المقرر"
        try:
            local_start = datetime.strptime(start_str, '%Y-%m-%dT%H:%M')
            local_end = datetime.strptime(end_str, '%Y-%m-%dT%H:%M')
            start_dt = local_start - timedelta(hours=LOCAL_TIMEZONE_OFFSET)
            end_dt = local_end - timedelta(hours=LOCAL_TIMEZONE_OFFSET)
            start_formatted = local_start.strftime('%Y-%m-%d الساعة %H:%M')
            end_formatted = local_end.strftime('الساعة %H:%M') if local_start.date() == local_end.date() else local_end.strftime('%Y-%m-%d الساعة %H:%M')
        except ValueError:
            start_formatted = start_str
            end_formatted = end_str
            start_dt = datetime.utcnow()
            end_dt = datetime.utcnow() + timedelta(hours=1)
        session_count = LiveSession.query.filter_by(course_id=course_id).count() + 1
        session_title = custom_title if custom_title else "جلسة رقم {} - {}".format(session_count, course_name)
        new_session = LiveSession(course_id=course_id, session_number=session_count, title=session_title, start_time=start_dt, end_time=end_dt, status='upcoming')
        db.session.add(new_session)
        message = "مواعيد جلسة مقرر ({}): من {} إلى {}".format(course_name, start_formatted, end_formatted)
        title = "موعد جلسة: {}".format(session_title)
        enrollments = StudentEnrollment.query.filter_by(course_id=course_id, is_approved=True).all()
        for e in enrollments:
            db.session.add(Announcement(sender_id=current_user.id, receiver_id=e.user_id, course_id=course_id, title=title, message=message, msg_type='session'))
        db.session.commit()
        flash('تم تحديد موعد الجلسة وإعلام الطلاب وإضافتها للأرشيف بنجاح', 'success')
    return redirect(url_for('virtual_classroom', course_id=course_id))

@app.route('/course/<int:course_id>/class/kick-student/<int:student_id>', methods=['POST'])
@teacher_required
def kick_student(course_id, student_id):
    return jsonify({'success': True})

@app.route('/course/<int:course_id>/chat/send', methods=['POST'])
@login_required
def send_chat(course_id):
    if current_user.is_guest:
        abort(403)
    if current_user.is_student and not check_student_enrolled(course_id):
        abort(403)
    message = request.form.get('message', '').strip()
    if message:
        db.session.add(ChatMessage(course_id=course_id, user_id=current_user.id, message=message))
        db.session.commit()
    return redirect(url_for('virtual_classroom', course_id=course_id) + '#chat-section')

@app.route('/course/<int:course_id>/forum/post', methods=['POST'])
@login_required
def add_forum_post(course_id):
    if current_user.is_guest:
        abort(403)
    if current_user.is_student and not check_student_enrolled(course_id):
        abort(403)
    title = request.form.get('title', '').strip()
    content = request.form.get('content', '').strip()
    post_type = request.form.get('post_type', 'discussion')
    if title and content:
        db.session.add(ForumPost(course_id=course_id, author_id=current_user.id, title=title, content=content, post_type=post_type))
        db.session.commit()
        flash('تم النشر', 'success')
    return redirect(url_for('virtual_classroom', course_id=course_id) + '#forum-section')

@app.route('/forum/post/<int:post_id>/reply', methods=['POST'])
@login_required
def reply_forum_post(post_id):
    post = ForumPost.query.get_or_404(post_id)
    if current_user.is_guest:
        abort(403)
    if current_user.is_student and not check_student_enrolled(post.course_id):
        abort(403)
    content = request.form.get('content', '').strip()
    if content:
        db.session.add(ForumReply(post_id=post.id, author_id=current_user.id, content=content))
        db.session.commit()
        flash('تم الرد', 'success')
    return redirect(url_for('virtual_classroom', course_id=post.course_id) + '#forum-section')

@app.route('/student/courses')
@student_required
def student_courses():
    lang = get_lang()
    logo = get_logo()
    student = current_user
    my_college = College.query.get(student.college_id) if student.college_id else None
    enrollments = StudentEnrollment.query.filter_by(user_id=student.id).all()
    enrolled_ids = set([e.course_id for e in enrollments if e.is_approved])
    pending_ids = set([e.course_id for e in enrollments if not e.is_approved])
    enrolled_courses = []
    seen_ids = set()
    for e in enrollments:
        if e.is_approved:
            course = Course.query.get(e.course_id)
            if course and course.id not in seen_ids:
                enrolled_courses.append(course)
                seen_ids.add(course.id)
    available_courses = []
    if my_college:
        for year in my_college.years:
            for sem in year.semesters:
                for course in sem.courses:
                    if course.id not in enrolled_ids and course.id not in seen_ids and course.is_active:
                        available_courses.append(course)
                        seen_ids.add(course.id)
        for dept in my_college.children:
            for year in dept.years:
                for sem in year.semesters:
                    for course in sem.courses:
                        if course.id not in enrolled_ids and course.id not in seen_ids and course.is_active:
                            available_courses.append(course)
                            seen_ids.add(course.id)
        if my_college.parent_id:
            parent = College.query.get(my_college.parent_id)
            if parent:
                for year in parent.years:
                    for sem in year.semesters:
                        for course in sem.courses:
                            if course.id not in enrolled_ids and course.id not in seen_ids and course.is_active:
                                available_courses.append(course)
                                seen_ids.add(course.id)
    req_courses = []
    req_unit = College.query.filter(College.name.like('%متطلبات%')).first()
    if req_unit:
        for year in req_unit.years:
            for sem in year.semesters:
                for course in sem.courses:
                    if course.id not in enrolled_ids and course.id not in seen_ids and course.is_active:
                        req_courses.append(course)
                        seen_ids.add(course.id)
    return render_template('student/courses.html', lang=lang, logo=logo, student=student, my_college=my_college, enrolled_courses=enrolled_courses, enrolled_ids=list(enrolled_ids), pending_ids=list(pending_ids), available_courses=available_courses, req_courses=req_courses, req_unit=req_unit)

@app.route('/student/dashboard')
@student_required
def student_dashboard():
    lang = get_lang()
    logo = get_logo()
    student = current_user
    enrollments = StudentEnrollment.query.filter_by(user_id=student.id, is_approved=True).all()
    courses = [e.course for e in enrollments]
    total_tests_taken = TestResult.query.filter_by(user_id=student.id).count()
    results = TestResult.query.filter_by(user_id=student.id).all()
    avg_score = sum(r.percentage for r in results) / len(results) if results else 0
    unread_msgs = Announcement.query.filter_by(receiver_id=student.id, is_read=False).count()
    rewards = StudentReward.query.filter_by(user_id=student.id).all()
    total_points = sum(r.points for r in rewards)
    return render_template('student/dashboard.html', lang=lang, logo=logo, student=student, courses=courses, total_tests_taken=total_tests_taken, avg_score=round(avg_score, 1), unread_msgs=unread_msgs, rewards=rewards, total_points=total_points)

@app.route('/student/course/<int:id>')
@login_required
def student_course_detail(id):
    lang = get_lang()
    logo = get_logo()
    course = Course.query.get_or_404(id)

    # التأكد من أن الطالب مسجل في هذا المقرر
    if current_user.is_student:
        enrollment = StudentEnrollment.query.filter_by(user_id=current_user.id, course_id=course.id, is_approved=True).first()
        if not enrollment:
            flash('أنت غير مسجل في هذا المقرر أو لم تتم الموافقة بعد', 'warning')
            return redirect(url_for('student_courses'))

    # جلب اختبارات الطالب التي قام بأدائها لمعرفة أي زر نعرض (أداء الاختبار أو تم الأداء)
    results = TestResult.query.filter_by(user_id=current_user.id).all()
    taken_test_ids = {r.test_id for r in results}

    return render_template('student/course_detail.html', lang=lang, logo=logo, course=course, taken_test_ids=taken_test_ids)

@app.route('/student/test/<int:id>/take', methods=['GET', 'POST'])
@student_required
def student_take_test(id):
    lang = get_lang()
    logo = get_logo()
    test = Test.query.get_or_404(id)

    if not check_student_enrolled(test.course_id):
        flash('عليك التسجيل والموافقة عليه أولاً لأداء الاختبار', 'warning')
        return redirect(url_for('student_courses'))

    existing_result = TestResult.query.filter_by(test_id=id, user_id=current_user.id).first()
    if existing_result:
        flash('لقد أجريت هذا الاختبار مسبقاً', 'warning')
        return redirect(url_for('student_tests'))

    if test.deadline and test.deadline < datetime.utcnow():
        flash('انتهى الموعد النهائي لهذا الاختبار', 'danger')
        return redirect(url_for('student_tests'))

    if not test.is_active:
        flash('هذا الاختبار غير متاح حالياً', 'warning')
        return redirect(url_for('student_tests'))

    questions = test.questions.order_by(Question.sort_order).all()
    if not questions:
        flash('لا توجد أسئلة في هذا الاختبار', 'warning')
        return redirect(url_for('student_tests'))

    if request.method == 'POST':
        answers = {}
        auto_score = 0
        auto_total = 0
        has_essay = False

        for q in questions:
            a = request.form.get('q_{}'.format(q.id), '').strip()
            answers[str(q.id)] = a

            if q.question_type in ['mcq', 'tf']:
                auto_total += q.marks
                if a == q.correct_answer:
                    auto_score += q.marks
            elif q.question_type == 'essay':
                has_essay = True
            else:
                auto_total += q.marks
                if a == q.correct_answer:
                    auto_score += q.marks

        total_all = sum(q.marks for q in questions)
        auto_pct = (auto_score / auto_total * 100) if auto_total > 0 else 0

        if has_essay:
            ai_eval = 'تقييم آلي: {}/{} للأسئلة التلقائية - الأسئلة المقالية بانتظار التصحيح اليدوي'.format(auto_score, auto_total)
        else:
            pct = (auto_score / total_all * 100) if total_all > 0 else 0
            ai_eval = 'تقييم AI: {}% - {} سؤال - {} صحيحة'.format(round(pct, 1), len(questions), int(auto_score))
            auto_pct = pct

        db.session.add(TestResult(
            test_id=test.id, user_id=current_user.id,
            score=auto_score, total_marks=total_all,
            percentage=round(auto_pct, 2),
            answers=json.dumps(answers, ensure_ascii=False),
            ai_evaluation=ai_eval
        ))
        db.session.add(StudentTracking(
            user_id=current_user.id, course_id=test.course_id,
            action_type='test_submit', action_details=test.name
        ))
        db.session.commit()

        if has_essay:
            flash('تم التسليم - علامتك الآلية: {}/{} (بانتظار تصحيح الأسئلة المقالية)'.format(auto_score, auto_total), 'info')
        else:
            flash('تم التسليم - درجتك: {}/{}'.format(auto_score, total_all), 'success')
        return redirect(url_for('student_tests'))

    return render_template('student/take_test.html', lang=lang, logo=logo, test=test, questions=questions)

@app.route('/student/results')
@student_required
def student_results():
    lang = get_lang()
    logo = get_logo()
    return render_template('student/results.html', lang=lang, logo=logo, results=TestResult.query.filter_by(user_id=current_user.id).order_by(TestResult.submitted_at.desc()).all())

@app.route('/student/assignments')
@student_required
def student_assignments():
    lang = get_lang()
    logo = get_logo()
    return render_template('student/assignments.html', lang=lang, logo=logo, announcements=Announcement.query.filter_by(receiver_id=current_user.id).order_by(Announcement.created_at.desc()).all())

@app.route('/student/messages')
@student_required
def student_messages():
    lang = get_lang()
    logo = get_logo()
    msgs = Announcement.query.filter_by(receiver_id=current_user.id).order_by(Announcement.created_at.desc()).all()
    Announcement.query.filter_by(receiver_id=current_user.id, is_read=False).update({Announcement.is_read: True})
    db.session.commit()
    return render_template('student/messages.html', lang=lang, logo=logo, messages=msgs)

@app.route('/student/rewards')
@student_required
def student_rewards():
    lang = get_lang()
    logo = get_logo()
    rewards = StudentReward.query.filter_by(user_id=current_user.id).order_by(StudentReward.created_at.desc()).all()
    return render_template('student/rewards.html', lang=lang, logo=logo, rewards=rewards, total_points=sum(r.points for r in rewards))

@app.route('/student/profile')
@student_required
def student_profile():
    lang = get_lang()
    logo = get_logo()
    student = current_user
    college = College.query.get(student.college_id) if student.college_id else None
    enrollments = StudentEnrollment.query.filter_by(user_id=student.id, is_approved=True).all()
    results = TestResult.query.filter_by(user_id=student.id).all()
    avg = sum(r.percentage for r in results) / len(results) if results else 0
    tracking = StudentTracking.query.filter_by(user_id=student.id).order_by(StudentTracking.created_at.desc()).limit(20).all()
    return render_template('student/profile.html', lang=lang, logo=logo, student=student, college=college, enrollments=enrollments, avg_score=round(avg, 1), tracking=tracking)

@app.route('/student/unenroll/<int:course_id>', methods=['POST'])
@student_required
def student_unenroll(course_id):
    enrollment = StudentEnrollment.query.filter_by(user_id=current_user.id, course_id=course_id).first()
    if enrollment:
        db.session.delete(enrollment)
        db.session.add(StudentTracking(user_id=current_user.id, course_id=course_id, action_type='unenroll', action_details='إلغاء تسجيل من مقرر'))
        db.session.commit()
        flash('تم إلغاء التسجيل من المقرر بنجاح', 'success')
    else:
        flash('أنت غير مسجل في هذا المقرر', 'warning')
    return redirect(url_for('student_courses'))

@app.route('/student/enroll/<int:course_id>', methods=['POST'])
@student_required
def student_enroll(course_id):
    enrollment = StudentEnrollment.query.filter_by(user_id=current_user.id, course_id=course_id).first()
    if enrollment:
        if enrollment.is_approved:
            flash('مسجل وموافق عليه مسبقاً', 'warning')
        else:
            flash('طلبك قيد المراجعة من إدارة المنصة', 'warning')
    else:
        new_enrollment = StudentEnrollment(user_id=current_user.id, course_id=course_id, is_approved=False)
        db.session.add(new_enrollment)
        db.session.commit()
        flash('تم تقديم طلب التسجيل بنجاح، بانتظار موافقة مدير المنصة', 'success')
    return redirect(url_for('student_course_detail', id=course_id))

@app.route('/archive')
@login_required
def academic_archive():
    lang = get_lang()
    logo = get_logo()
    available_years = [r[0].strftime('%Y') for r in db.session.query(Course.created_at).distinct().all() if r[0]]
    available_years = sorted(list(set(available_years)), reverse=True)
    available_semesters = [('first', 'الفصل الدراسي الأول'), ('second', 'الفصل الدراسي الثاني'), ('summer', 'الفصل الدراسي الصيفي')]
    all_cols = College.query.filter_by(is_active=True).order_by(College.sort_order).all()
    college_tree = []
    for c in all_cols:
        if c.parent_id is None:
            college_tree.append({'college': c, 'departments': [d for d in all_cols if d.parent_id == c.id]})
    accessible_college_tree = []
    if current_user.is_admin or current_user.is_guest:
        accessible_college_tree = college_tree
    else:
        if current_user.is_teacher:
            c_ids = [c.id for c in Course.query.filter_by(teacher_id=current_user.id).all()]
        else:
            c_ids = [e.course_id for e in StudentEnrollment.query.filter_by(user_id=current_user.id, is_approved=True).all()]
        valid_college_ids = set()
        for cid in c_ids:
            course = Course.query.get(cid)
            if course and course.semester and course.semester.year and course.semester.year.college:
                col = course.semester.year.college
                valid_college_ids.add(col.id)
                if col.parent_id:
                    valid_college_ids.add(col.parent_id)
        for item in college_tree:
            if item['college'].id in valid_college_ids:
                accessible_college_tree.append(item)
            else:
                valid_depts = [d for d in item['departments'] if d.id in valid_college_ids]
                if valid_depts:
                    accessible_college_tree.append({'college': item['college'], 'departments': valid_depts})
    selected_year = request.args.get('year', '')
    selected_college_id = request.args.get('college_id', 0, type=int)
    selected_semester = request.args.get('semester', '')
    archived_data = []
    if selected_year:
        base_course_q = Course.query.filter(db.func.strftime('%Y', Course.created_at) == selected_year)
        if current_user.is_teacher:
            base_course_q = base_course_q.filter(Course.teacher_id == current_user.id)
        elif current_user.is_student:
            en_ids = [e.course_id for e in StudentEnrollment.query.filter_by(user_id=current_user.id, is_approved=True).all()]
            base_course_q = base_course_q.filter(Course.id.in_(en_ids))
        if selected_college_id > 0:
            target_col = College.query.get(selected_college_id)
            if target_col:
                col_ids = [target_col.id] + [d.id for d in target_col.children.all()] if target_col.parent_id is None else [target_col.id]
                yr_ids = [y.id for y in StudyYear.query.filter(StudyYear.college_id.in_(col_ids)).all()]
                sem_ids = [s.id for s in Semester.query.filter(Semester.year_id.in_(yr_ids)).all()]
                base_course_q = base_course_q.filter(Course.semester_id.in_(sem_ids))
        if selected_semester:
            sem_ids = [s.id for s in Semester.query.filter_by(semester_type=selected_semester).all()]
            base_course_q = base_course_q.filter(Course.semester_id.in_(sem_ids))
        valid_courses = base_course_q.all()
        for course in valid_courses:
            lecs = Lecture.query.filter(Lecture.course_id == course.id).all()
            tests = Test.query.filter(Test.course_id == course.id).all()
            sess = LiveSession.query.filter(LiveSession.course_id == course.id).all()
            students_data = []
            passed, failed, no_result = 0, 0, 0
            enrollments = StudentEnrollment.query.filter_by(course_id=course.id, is_approved=True).all()
            for en in enrollments:
                results = TestResult.query.filter_by(user_id=en.user_id).join(Test).filter(Test.course_id == course.id).all()
                avg = sum(r.percentage for r in results) / len(results) if results else 0
                status = 'ناجح' if avg >= 60 else ('راسب' if results else 'لم يختبر')
                if status == 'ناجح':
                    passed += 1
                elif status == 'راسب':
                    failed += 1
                else:
                    no_result += 1
                students_data.append({'user': en.user, 'avg': round(avg, 1), 'status': status})
            archived_data.append({'course': course, 'lectures': lecs, 'tests': tests, 'sessions': sess, 'students': students_data, 'passed': passed, 'failed': failed, 'no_result': no_result})
    return render_template('archive.html', lang=lang, logo=logo, available_years=available_years, available_semesters=available_semesters, accessible_college_tree=accessible_college_tree, selected_year=selected_year, selected_college_id=selected_college_id, selected_semester=selected_semester, archived_data=archived_data)

def create_app():
    return app
@app.route('/student/tests')
@login_required
def student_tests():
    lang = get_lang()
    logo = get_logo()

    # التأكد من أن المستخدم طالب
    if not current_user.is_student:
        flash('هذه الصفحة مخصصة للطلاب فقط', 'warning')
        return redirect(url_for('index'))

    # جلب المقررات المسجل بها الطالب والموافق عليها
    enrollments = StudentEnrollment.query.filter_by(user_id=current_user.id, is_approved=True).all()
    courses = [e.course for e in enrollments if e.course]
    course_ids = [c.id for c in courses]

    # جلب الاختبارات التابعة لمقررات الطالب
    if not course_ids:
        tests = []
    else:
        tests = Test.query.filter(Test.course_id.in_(course_ids), Test.is_active == True).order_by(Test.deadline.asc()).all()

    # جلب الاختبارات التي أنجزها الطالب مسبقاً
    results = TestResult.query.filter_by(user_id=current_user.id).all()
    taken_test_ids = {r.test_id for r in results}

    return render_template('student/tests.html', lang=lang, logo=logo, tests=tests, taken_test_ids=taken_test_ids)
def initialize_platform():
    with app.app_context():
        db.create_all()
        # إنشاء حساب المدير تلقائياً إذا لم يكن موجوداً (مهم جداً للنشر السحابي)
        if not User.query.filter_by(username='admin').first():
            admin = User(first_name='مدير', second_name='المنصة', username='admin', user_type='admin', is_active=True)
            admin.set_password('admin123')
            db.session.add(admin)
            db.session.commit()
            print("Admin user created successfully!")

if __name__ == '__main__':
    initialize_platform()
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
else:
    initialize_platform()# ===== API لجدول المستخدمين (DataTables Server-side) =====
@app.route('/admin/users/data')
@admin_required
def admin_users_data():
    draw = request.args.get('draw', 1, type=int)
    start = request.args.get('start', 0, type=int)
    length = request.args.get('length', 10, type=int)
    search = request.args.get('search[value]', '', type=str)
    user_type_filter = request.args.get('type', '', type=str)

    query = User.query
    if user_type_filter:
        query = query.filter_by(user_type=user_type_filter)
    if search:
        query = query.filter(db.or_(
            User.first_name.contains(search),
            User.second_name.contains(search),
            User.username.contains(search),
            User.mobile.contains(search)
        ))

    total_records = User.query.count()
    filtered_records = query.count()

    users = query.order_by(User.created_at.desc()).offset(start).limit(length).all()
    data = []
    for u in users:
        status_html = '<span class="text-success"><i class="fas fa-check-circle"></i> نشط</span>' if u.is_active else '<span class="text-danger"><i class="fas fa-times-circle"></i> معطل</span>'
        actions_html = f'''<div class="action-buttons">
            <a href="/admin/users/edit/{u.id}" class="btn btn-sm btn-info btn-3d"><i class="fas fa-edit"></i></a>
            <form action="/admin/users/delete/{u.id}" method="POST" style="display:inline;" onsubmit="return confirm('هل أنت متأكد؟')">
                <input type="hidden" name="csrf_token" value="{csrf_token()}">
                <button type="submit" class="btn btn-sm btn-danger btn-3d"><i class="fas fa-trash"></i></button>
            </form>
        </div>'''
        type_map = {'admin': 'مدير', 'teacher': 'مدرس', 'student': 'طالب', 'guest': 'ضيف'}
        data.append([
            '', # سيتم ترقيمه تلقائياً في الجدول
            f'{u.first_name} {u.second_name}',
            u.username,
            type_map.get(u.user_type, u.user_type),
            u.number or '-',
            u.mobile or '-',
            status_html,
            actions_html
        ])

    return jsonify({
        "draw": draw,
        "recordsTotal": total_records,
        "recordsFiltered": filtered_records,
        "data": data
    })
